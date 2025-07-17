from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Sum
from django.utils import timezone
from datetime import date, timedelta
from django.core.paginator import Paginator
from django.views.decorators.http import require_http_methods
import json
from .models import User, Project, Task, WorkEntry
from .forms import *
from io import BytesIO
import calendar
# work hour calculations of employees and projects 
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Q
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.utils import timezone
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Project, WorkEntry, Task
from django.contrib.auth import get_user_model
from decimal import Decimal
from django.db.models import Case, When, Value, CharField
# Admin Views


from django.template.loader import render_to_string
from django.views.decorators.csrf import csrf_protect
from django.core.paginator import Paginator
import json
from django.views.decorators.csrf import csrf_exempt


def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('email')
        password = request.POST.get('password')
        user = authenticate(request, email=username, password=password)
        
        if user is not None and user.is_active:
            login(request, user)
            if user.role == 'admin':
                return redirect('admin_dashboard')
            else:
                return redirect('employee_dashboard')
        else:
            messages.error(request, 'Invalid username or password')
    
    return render(request, 'login.html')

@login_required
def logout_view(request):
    logout(request)
    return redirect('login_view')


@login_required
def admin_dashboard(request):
    if request.user.role != 'admin':
        return redirect('employee_dashboard')
    
    # Dashboard statistics
    total_employees = User.objects.filter(role='user', is_active=True).count()
    total_projects = Project.objects.filter(status='active').count()
    total_tasks = Task.objects.count()
    
    # Recent work entries
    recent_entries = WorkEntry.objects.select_related('employee', 'project').order_by('-created_at')[:5]
    
    # Active projects
    active_projects = Project.objects.filter(status='active').order_by('-created_at')[:5]
    
    context = {
        'total_employees': total_employees,
        'total_projects': total_projects,
        'total_tasks': total_tasks,
        'recent_entries': recent_entries,
        'active_projects': active_projects,
    }
    return render(request, 'admin/dashboard.html', context)

@login_required
def employee_dashboard(request):
    if request.user.role == 'admin':
        return redirect('admin_dashboard')
    
    # Employee's active projects
    active_projects = Project.objects.filter(status='active')[:5]
    
    # Employee's recent work entries
    recent_entries = WorkEntry.objects.filter(employee=request.user).order_by('-work_date')[:5]
    
    # Employee's assigned tasks
    assigned_tasks = Task.objects.filter(assigned_to=request.user).order_by('-created_at')[:5]
    
    context = {
        'active_projects': active_projects,
        'recent_entries': recent_entries,
        'assigned_tasks': assigned_tasks,
    }
    return render(request, 'employee/dashboard.html', context)





@login_required
def employee_management(request):
    """Display employee management page with all employees"""
    if request.user.role != 'admin':
        return redirect('employee_dashboard')
    
    employees = User.objects.filter(role='user').order_by('-date_joined')
    form = CustomUserForm()  # For the add employee modal
    
    context = {
        'employees': employees,
        'form': form,
    }
    return render(request, 'admin/employee_management.html', context)


@login_required
def add_employee(request):
    """Add new employee"""
    if request.user.role != 'admin':
        return redirect('employee_dashboard')
    
    if request.method == 'POST':
        form = CustomUserForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save(commit=False)
            user.role = 'user'
            user.save()
            messages.success(request, 'Employee added successfully!')
            return redirect('employee_management')
        else:
            # Return form errors for AJAX handling
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                })
            messages.error(request, 'Please correct the errors below.')
    else:
        form = CustomUserForm()
    
    return render(request, 'admin/employee_management.html', {'form': form})


@login_required
@csrf_exempt
def edit_employee(request, user_id):
    """Edit employee details"""
    if request.user.role != 'admin':
        return redirect('employee_dashboard')
    
    employee = get_object_or_404(User, id=user_id, role='user')
    
    if request.method == 'POST':
        form = CustomUserForm(request.POST, request.FILES, instance=employee)
        if form.is_valid():
            form.save()
            
            # Return JSON response for AJAX requests
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Employee updated successfully!'
                })
            
            messages.success(request, 'Employee updated successfully!')
            return redirect('employee_management')
        else:
            # Return form errors for AJAX handling
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                })
            messages.error(request, 'Please correct the errors below.')
    else:
        form = CustomUserForm(instance=employee)
    
    # For AJAX requests, return the form HTML
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        form_html = render_to_string('admin/employee_edit_form.html', {
            'form': form,
            'employee': employee
        })
        return JsonResponse({'form_html': form_html})
    
    return render(request, 'admin/edit_employee.html', {
        'form': form,
        'employee': employee
    })


@login_required
def view_employee(request, user_id):
    """View employee details"""
    if request.user.role != 'admin':
        return redirect('employee_dashboard')
    
    employee = get_object_or_404(User, id=user_id, role='user')
    
    # For AJAX requests, return the employee details HTML
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        details_html = render_to_string('admin/employee_details.html', {
            'employee': employee
        })
        return JsonResponse({'details_html': details_html})
    
    return render(request, 'admin/employee_details.html', {'employee': employee})


@login_required
@csrf_protect
def toggle_employee_status(request, user_id):
    """Toggle employee active/inactive status"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'message': 'Unauthorized'})
    
    if request.method == 'POST':
        employee = get_object_or_404(User, id=user_id, role='user')
        employee.is_active = not employee.is_active
        employee.save()
        
        status = 'activated' if employee.is_active else 'deactivated'
        return JsonResponse({
            'success': True,
            'message': f'Employee {status} successfully!'
        })
    
    return JsonResponse({'success': False, 'message': 'Invalid request'})


# views.py
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Q
from django.utils import timezone
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from decimal import Decimal
from django.contrib.auth import get_user_model

User = get_user_model()

@login_required
def employee_work_entries(request, employee_id):
    employee = get_object_or_404(User, id=employee_id)
    
    # Get date range from request (default to current month)
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    if not from_date:
        from_date = timezone.now().replace(day=1).date()
    else:
        from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
    
    if not to_date:
        to_date = timezone.now().date()
    else:
        to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
    
    # Get work entries for the employee within date range
    work_entries = WorkEntry.objects.filter(
        employee=employee,
        work_date__range=[from_date, to_date]
    ).select_related('project').order_by('-work_date')
    
    # Calculate summary statistics
    total_working_hours = work_entries.aggregate(
        total_hours=Sum('working_hours')
    )['total_hours'] or Decimal('0.00')
    
    # Calculate manhour cost (employee manhour rate * working hours)
    employee_manhour_rate = employee.man_hour_of_employee or Decimal('0.00')
    total_manhour_cost = total_working_hours * employee_manhour_rate
    
    # Project-wise breakdown
    project_summary = {}
    for entry in work_entries:
        project_id = entry.project.id
        if project_id not in project_summary:
            project_summary[project_id] = {
                'project': entry.project,
                'total_hours': Decimal('0.00'),
                'total_cost': Decimal('0.00'),
                'entries': []
            }
        
        project_summary[project_id]['total_hours'] += entry.working_hours
        project_summary[project_id]['total_cost'] += (entry.working_hours * employee_manhour_rate)
        project_summary[project_id]['entries'].append(entry)
    
    # Convert to list for template
    project_data = list(project_summary.values())
    
    # Calculate average daily hours
    total_days = (to_date - from_date).days + 1
    working_days = work_entries.values('work_date').distinct().count()
    avg_daily_hours = total_working_hours / working_days if working_days > 0 else Decimal('0.00')
    
    context = {
        'employee': employee,
        'work_entries': work_entries,
        'project_data': project_data,
        'from_date': from_date,
        'to_date': to_date,
        'total_working_hours': total_working_hours,
        'employee_manhour_rate': employee_manhour_rate,
        'total_manhour_cost': total_manhour_cost,
        'working_days': working_days,
        'avg_daily_hours': avg_daily_hours,
        'total_projects': len(project_data),
    }
    
    return render(request, 'admin/employee_work_entries.html', context)

@login_required
def download_employee_work_excel(request, employee_id):
    employee = get_object_or_404(User, id=employee_id)
    
    # Get date range from request
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    if not from_date:
        from_date = timezone.now().replace(day=1).date()
    else:
        from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
    
    if not to_date:
        to_date = timezone.now().date()
    else:
        to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
    
    # Get work entries
    work_entries = WorkEntry.objects.filter(
        employee=employee,
        work_date__range=[from_date, to_date]
    ).select_related('project').order_by('-work_date')
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{employee.first_name} Work Report"
    
    # Define styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    sub_header_font = Font(bold=True)
    sub_header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    # Employee Information Section
    ws['A1'] = 'EMPLOYEE WORK REPORT'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:H1')
    
    ws['A3'] = 'Employee Information'
    ws['A3'].font = sub_header_font
    ws['A3'].fill = sub_header_fill
    ws.merge_cells('A3:H3')
    
    employee_info = [
        ['Employee Name:', f"{employee.first_name} {employee.last_name}"],
        ['Employee ID:', str(employee.id)],
        ['Email:', employee.email],
        ['Designation:', employee.designation or 'N/A'],
        ['Man Hour Rate (USD):', f"${employee.man_hour_of_employee:.2f}"],
        ['Report Period:', f"{from_date} to {to_date}"],
    ]
    
    row = 4
    for info in employee_info:
        ws[f'A{row}'] = info[0]
        ws[f'B{row}'] = info[1]
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Summary Section
    row += 1
    ws[f'A{row}'] = 'SUMMARY'
    ws[f'A{row}'].font = sub_header_font
    ws[f'A{row}'].fill = sub_header_fill
    ws.merge_cells(f'A{row}:H{row}')
    
    # Calculate totals
    total_hours = work_entries.aggregate(Sum('working_hours'))['working_hours__sum'] or Decimal('0.00')
    total_cost = total_hours * employee.man_hour_of_employee
    working_days = work_entries.values('work_date').distinct().count()
    
    summary_data = [
        ['Total Working Hours:', f"{total_hours:.2f} hours"],
        ['Total Cost (USD):', f"${total_cost:.2f}"],
        ['Working Days:', str(working_days)],
        ['Average Hours/Day:', f"{(total_hours/working_days if working_days > 0 else 0):.2f} hours"],
    ]
    
    row += 1
    for summary in summary_data:
        ws[f'A{row}'] = summary[0]
        ws[f'B{row}'] = summary[1]
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Project-wise Summary
    row += 2
    ws[f'A{row}'] = 'PROJECT-WISE SUMMARY'
    ws[f'A{row}'].font = sub_header_font
    ws[f'A{row}'].fill = sub_header_fill
    ws.merge_cells(f'A{row}:H{row}')
    
    row += 1
    project_headers = ['Project ID', 'Project Name', 'Total Hours', 'Cost (USD)']
    for col, header in enumerate(project_headers, 1):
        ws.cell(row=row, column=col, value=header)
        ws.cell(row=row, column=col).font = header_font
        ws.cell(row=row, column=col).fill = header_fill
    
    # Calculate project-wise data
    project_summary = {}
    for entry in work_entries:
        project_id = entry.project.project_id
        if project_id not in project_summary:
            project_summary[project_id] = {
                'name': entry.project.name,
                'hours': Decimal('0.00'),
                'cost': Decimal('0.00')
            }
        project_summary[project_id]['hours'] += entry.working_hours
        project_summary[project_id]['cost'] += (entry.working_hours * employee.man_hour_of_employee)
    
    row += 1
    for project_id, data in project_summary.items():
        ws[f'A{row}'] = project_id
        ws[f'B{row}'] = data['name']
        ws[f'C{row}'] = f"{data['hours']:.2f}"
        ws[f'D{row}'] = f"${data['cost']:.2f}"
        row += 1
    
    # Detailed Work Entries
    row += 2
    ws[f'A{row}'] = 'DETAILED WORK ENTRIES'
    ws[f'A{row}'].font = sub_header_font
    ws[f'A{row}'].fill = sub_header_fill
    ws.merge_cells(f'A{row}:H{row}')
    
    row += 1
    headers = ['Date', 'Project ID', 'Project Name', 'Start Time', 'End Time', 'Hours', 'Rate (USD)', 'Cost (USD)', 'Description']
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
        ws.cell(row=row, column=col).font = header_font
        ws.cell(row=row, column=col).fill = header_fill
    
    # Add work entries data
    row += 1
    for entry in work_entries:
        entry_cost = entry.working_hours * employee.man_hour_of_employee
        ws[f'A{row}'] = entry.work_date.strftime('%Y-%m-%d')
        ws[f'B{row}'] = entry.project.project_id
        ws[f'C{row}'] = entry.project.name
        ws[f'D{row}'] = entry.start_time.strftime('%H:%M')
        ws[f'E{row}'] = entry.end_time.strftime('%H:%M')
        ws[f'F{row}'] = f"{entry.working_hours:.2f}"
        ws[f'G{row}'] = f"${employee.man_hour_of_employee:.2f}"
        ws[f'H{row}'] = f"${entry_cost:.2f}"
        ws[f'I{row}'] = entry.description
        row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{employee.first_name}_{employee.last_name}_work_report_{from_date}_{to_date}.xlsx"'
    
    wb.save(response)
    return response




@login_required
@csrf_protect
def delete_employee(request, user_id):
    """Delete employee"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'message': 'Unauthorized'})
    
    if request.method == 'POST':
        employee = get_object_or_404(User, id=user_id, role='user')
        employee_name = f"{employee.first_name} {employee.last_name}"
        employee.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Employee {employee_name} deleted successfully!'
        })
    
    return JsonResponse({'success': False, 'message': 'Invalid request'})


@login_required
def employee_search(request):
    """Search employees with filters"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'message': 'Unauthorized'})
    
    query = request.GET.get('q', '')
    role_filter = request.GET.get('role', '')
    status_filter = request.GET.get('status', '')
    
    employees = User.objects.filter(role='user')
    
    # Apply search filter
    if query:
        employees = employees.filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query) |
            Q(phone_number__icontains=query) |
            Q(designation__icontains=query)
        )
    
    # Apply role filter
    if role_filter:
        employees = employees.filter(role=role_filter)
    
    # Apply status filter
    if status_filter:
        is_active = status_filter == 'active'
        employees = employees.filter(is_active=is_active)
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(employees, 10)
    employees_page = paginator.get_page(page)
    
    # Return JSON response for AJAX requests
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        employees_data = []
        for employee in employees_page:
            employees_data.append({
                'id': employee.id,
                'name': f"{employee.first_name} {employee.last_name or ''}",
                'email': employee.email,
                'phone': employee.phone_number or '-',
                'designation': employee.designation or '-',
                'role': employee.role,
                'is_active': employee.is_active,
                'date_joined': employee.date_joined.strftime('%b %d, %Y'),
                'profile_picture': employee.profile_picture.url if employee.profile_picture else None
            })
        
        return JsonResponse({
            'success': True,
            'employees': employees_data,
            'total_pages': paginator.num_pages,
            'current_page': employees_page.number,
            'total_count': paginator.count
        })
    
    return render(request, 'admin/employee_management.html', {
        'employees': employees_page,
        'query': query,
        'role_filter': role_filter,
        'status_filter': status_filter
    })

@login_required
def project_management(request):
    if request.user.role != 'admin':
        return redirect('employee_dashboard')
    
    projects = Project.objects.all().order_by('-created_at')
    paginator = Paginator(projects, 10)
    page = request.GET.get('page')
    projects = paginator.get_page(page)
    form = ProjectForm()
    
    return render(request, 'admin/project_management.html', {'projects': projects, 'form':form})

@login_required
def add_project(request):
    if request.user.role != 'admin':
        return redirect('employee_dashboard')
    
    if request.method == 'POST':
        form = ProjectForm(request.POST)
        if form.is_valid():
            project = form.save(commit=False)
            project.created_by = request.user
            project.save()
            messages.success(request, 'Project created successfully!')
            return redirect('project_management')
    else:
        form = ProjectForm()
    
    return render(request, 'admin/add_project.html', {'form': form})

@login_required
def edit_project(request, project_id):
    if request.user.role != 'admin':
        return redirect('employee_dashboard')
    
    project = get_object_or_404(Project, id=project_id)
    
    if request.method == 'POST':
        form = ProjectForm(request.POST, instance=project)
        if form.is_valid():
            form.save()
            messages.success(request, 'Project updated successfully!')
            return redirect('project_management')
    else:
        form = ProjectForm(instance=project)
    
    return render(request, 'admin/edit_project.html', {'form': form, 'project': project})

@login_required
def delete_project(request, project_id):
    if request.user.role != 'admin':
        return redirect('employee_dashboard')
    
    project = get_object_or_404(Project, id=project_id)
    project_name = project.name
    project.delete()
    messages.success(request, f'Project "{project_name}" deleted successfully!')
    return redirect('project_management')

@login_required
def project_details(request, project_id):
    if request.user.role != 'admin':
        return redirect('employee_dashboard')
    
    project = get_object_or_404(Project, id=project_id)
    tasks = Task.objects.filter(project=project)
    work_entries = WorkEntry.objects.filter(project=project)
    employees =  CustomUser.objects.filter(role = "user", is_active = True)
    return render(request, 'admin/project_details.html', {
        'project': project,
        'tasks': tasks,
        'work_entries': work_entries,
        'employees':employees
    })

@login_required
def close_project(request, project_id):
    if request.user.role != 'admin':
        return redirect('employee_dashboard')
    
    project = get_object_or_404(Project, id=project_id)
    if project.status == "active":
        project.status = 'closed'

        messages.success(request, f'Project "{project.name}" has been closed.')

    else:
        project.status = "active"
        messages.success(request, f'Project "{project.name}" has been Re Opened.')

    project.save()
    return redirect('project_management')

@login_required
def work_entry_management(request):
    if request.user.role != 'admin':
        return redirect('employee_dashboard')
    
    work_entries = WorkEntry.objects.select_related('employee', 'project').order_by('-work_date')
    
    # Filters
    employee_filter = request.GET.get('employee')
    project_filter = request.GET.get('project')
    date_filter = request.GET.get('date')
    
    if employee_filter:
        work_entries = work_entries.filter(employee_id=employee_filter)
    if project_filter:
        work_entries = work_entries.filter(project_id=project_filter)
    if date_filter:
        work_entries = work_entries.filter(work_date=date_filter)
    
    paginator = Paginator(work_entries, 20)
    page = request.GET.get('page')
    work_entries = paginator.get_page(page)
    
    employees = User.objects.filter(role='user', is_active=True)
    projects = Project.objects.all()
  
    try:
        if request.method == "POST":
            form = WorkEntryFormAdmin(request.POST)
            if form.is_valid():
                instance = form.save()
                instance.save()
                messages.success(request, 'Work entry was created')
                return redirect("work_entry_management")
            else:
                messages.info(request, form.errors.as_text())
                return redirect("work_entry_management")

    except Exception as e:
        messages.info(request, f"Something went wrong: {str(e)}")
        return redirect("work_entry_management")


    context = {
        'work_entries': work_entries,
        'employees': employees,
        'projects': projects,
        "form":WorkEntryFormAdmin()
    }
    return render(request, 'admin/work_entry_management.html', context)

@login_required
def edit_work_entry(request, entry_id):
    if request.user.role != 'admin':
        return redirect('employee_dashboard')
    
    work_entry = get_object_or_404(WorkEntry, id=entry_id)
    
    if request.method == 'POST':
        form = WorkEntryFormAdmin(request.POST, instance=work_entry)
        if form.is_valid():
            entry = form.save(commit=False)
            entry.updated_by = request.user
            entry.save()
            messages.success(request, 'Work entry updated successfully!')
            return redirect('work_entry_management')
    else:
        form = WorkEntryFormAdmin(instance=work_entry)
    
    return render(request, 'admin/edit_work_entry.html', {'form': form, 'work_entry': work_entry})


@login_required
def edit_work_entry_employee(request, entry_id):
    if request.user.role != 'user':
        return redirect('admin_dashboard')
    
    work_entry = get_object_or_404(WorkEntry, id=entry_id)
    
    if request.method == 'POST':
        form = WorkEntryForm(request.POST, instance=work_entry)
        if form.is_valid():
            entry = form.save(commit=False)
            entry.updated_by = request.user
            entry.save()
            messages.success(request, 'Work entry updated successfully!')
            return redirect('work_entry_management')
    else:
        form = WorkEntryForm(instance=work_entry)
    
    return render(request, 'employee/edit_work_entry.html', {'form': form, 'work_entry': work_entry})

from datetime import datetime

@login_required
def add_work_entry_admin(request, project_id):

    # projects = Project.objects.filter(status = "active")
    project = get_object_or_404(Project,id = project_id)
    if request.user.role != 'admin':
        return redirect('employee_dashboard')
    
    try:
        if request.method == 'POST':
            employee_id = request.POST.get("employee")
            work_date_str = request.POST.get("work_date")
            start_time_str = request.POST.get("start_time")
            end_time_str = request.POST.get("end_time")
            working_hours = request.POST.get("working_hours")
            description = request.POST.get("description")

            try:
                work_date = datetime.strptime(work_date_str, "%Y-%m-%d").date()
                start_time = datetime.strptime(start_time_str, "%H:%M").time()
                end_time = datetime.strptime(end_time_str, "%H:%M").time()
            except ValueError:
                messages.error(request, "Invalid date or time format.")
                return redirect('work_entry_management')

            if start_time >= end_time:
                messages.info(request, 'End time must be later than start time.')
                return redirect('work_entry_management')

            employee = get_object_or_404(CustomUser, id=employee_id)

            WorkEntry.objects.create(
                employee=employee,
                project=project,
                work_date=work_date,
                start_time=start_time,
                end_time=end_time,
                working_hours=working_hours,
                description=description
            )

            messages.success(request, 'Work entry added successfully!')
            return redirect('work_entry_management')
    except Exception as e:
        messages.info(request, f"Something went wrong: {str(e)}")
        return redirect('work_entry_management')
        

    else:
        form = WorkEntryFormAdmin()
    return redirect('work_entry_management')
    

# Employee Views
@login_required
def add_work_entry(request):
    projects = Project.objects.filter(status = "active")
    if request.user.role != 'user':
        return redirect('admin_dashboard')
    
    if request.method == 'POST':
        form = WorkEntryForm(request.POST)
        if form.is_valid():
            work_entry = form.save(commit=False)
            work_entry.employee = request.user
            
            # Check if work date is within allowed range
            today = date.today()
            if (today - work_entry.work_date).days > 2:
                messages.error(request, 'You can only add work entries for today or yesterday.')
                return redirect('my_work_entries')
            
            work_entry.save()
            messages.success(request, 'Work entry added successfully!')
            return redirect('my_work_entries')
        else:
            errors = []
            for field, error_list in form.errors.items():
                for error in error_list:
                    errors.append(f" {error}")



            messages.info(request, " | ".join(errors))
            return redirect('my_work_entries')

    else:
        form = WorkEntryForm()
    
    return render(request, 'employee/add_work_entry.html', {'form': form, "projects":projects})

@login_required
def my_work_entries(request):
    if request.user.role != 'user':
        return redirect('admin_dashboard')
    
    # Base queryset
    work_entries = WorkEntry.objects.filter(employee=request.user).select_related('project').order_by('-work_date', '-created_at')
    
    # Filters
    project_filter = request.GET.get('project')
    date_filter = request.GET.get('date')
    
    if project_filter:
        work_entries = work_entries.filter(project_id=project_filter)
        
    if date_filter:
        work_entries = work_entries.filter(work_date=date_filter)
    
    # Pagination
    paginator = Paginator(work_entries, 20)
    page = request.GET.get('page')
    work_entries = paginator.get_page(page)
    
    # Get active projects for filter dropdown
    projects = Project.objects.filter(status='active').order_by('name')
    
    # Calculate summary stats for the current filter
    total_entries = work_entries.paginator.count if work_entries else 0
    total_hours = WorkEntry.objects.filter(employee=request.user).aggregate(
        total=Sum('working_hours')
    )['total'] or 0
    
    # Get recent activity summary
    today = date.today()
    recent_entries = WorkEntry.objects.filter(
        employee=request.user,
        work_date__gte=today - timedelta(days=7)
    ).count()
    
    context = {
        'work_entries': work_entries,
        'projects': projects,
        'total_entries': total_entries,
        'total_hours': total_hours,
        'recent_entries': recent_entries,
        'current_filters': {
            'project': project_filter,
            'date': date_filter,
        },
        'today': today,
        'form':WorkEntryForm()
    }
    
    return render(request, 'employee/my_work_entries.html', context)

@login_required
def view_work_entry(request, entry_id):
    """View for viewing work entry details"""
    if request.user.role != 'user':
        return redirect('admin_dashboard')
    
    try:
        entry = WorkEntry.objects.get(id=entry_id, employee=request.user)
    except WorkEntry.DoesNotExist:
        messages.error(request, 'Work entry not found.')
        return redirect('my_work_entries')
    
    return render(request, 'employee/view_work_entry.html', {'entry': entry})

@login_required
def view_work_entry_admin(request, entry_id):
    """View for viewing work entry details"""
    if request.user.role != 'admin':
        return redirect('employee_dashboard')
    
    try:
        entry = WorkEntry.objects.get(id=entry_id)
    except WorkEntry.DoesNotExist:
        messages.info(request, 'Work entry not found.')
        return redirect('work_entry_management')
    
    return render(request, 'admin/view_work_entry.html', {'entry': entry})





# API Views
@login_required
def get_available_dates(request):
    """Return available dates for work entry (today and yesterday)"""
    today = date.today()
    yesterday = today - timedelta(days=1)
    
    dates = [
        {'value': today.isoformat(), 'label': f'Today ({today.strftime("%B %d, %Y")})'},
        {'value': yesterday.isoformat(), 'label': f'Yesterday ({yesterday.strftime("%B %d, %Y")})'},
    ]
    
    return JsonResponse({'dates': dates})

@login_required
def get_active_projects(request):
    """Return active projects for dropdown"""
    projects = Project.objects.filter(status='active').values('id', 'name')
    return JsonResponse({'projects': list(projects)})







User = get_user_model()

@login_required
def project_work_hours_report(request, project_id):
    """View to display work hours and payment calculation for a specific project"""
    project = get_object_or_404(Project, id=project_id)
    
    # Get date filters from request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    employee_id = request.GET.get('employee')
    
    # Base queryset
    work_entries = WorkEntry.objects.filter(project=project)
    
    # Apply date filters
    if start_date:
        work_entries = work_entries.filter(work_date__gte=start_date)
    if end_date:
        work_entries = work_entries.filter(work_date__lte=end_date)
    if employee_id:
        work_entries = work_entries.filter(employee_id=employee_id)
    
    # Calculate employee-wise statistics
    employee_stats = work_entries.values(
        'employee__id',
        'employee__first_name',
        'employee__last_name',
        'employee__man_hour_of_employee',
        'employee__designation'
    ).annotate(
        total_hours=Sum('working_hours'),
        total_entries=Count('id')
    ).order_by('employee__first_name')
    
    # Calculate payments
    for stat in employee_stats:
        total_hours = stat['total_hours'] or 0
        hourly_rate = stat['employee__man_hour_of_employee'] or 0
        stat['total_payment'] = total_hours * hourly_rate
        stat['employee_name'] = f"{stat['employee__first_name']} {stat['employee__last_name'] or ''}"
    
    # Calculate project totals
    project_totals = work_entries.aggregate(
        total_hours=Sum('working_hours'),
        total_entries=Count('id')
    )
    
    # Calculate total payment for project
    total_payment = sum(stat['total_payment'] for stat in employee_stats)
    
    # Get all employees for filter dropdown
    employees = User.objects.filter(work_entries__project=project).distinct()
    
    # Recent work entries for detailed view
    recent_entries = work_entries.select_related('employee').order_by('-work_date', '-created_at')[:20]
    
    context = {
        'project': project,
        'employee_stats': employee_stats,
        'project_totals': project_totals,
        'total_payment': total_payment,
        'employees': employees,
        'recent_entries': recent_entries,
        'start_date': start_date,
        'end_date': end_date,
        'selected_employee': employee_id,
    }
    
    return render(request, 'admin/project_work_hours_report.html', context)

@login_required
def all_projects_work_summary(request):
    """View to display work hours summary for all projects"""
    
    # Get date filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Base queryset
    work_entries = WorkEntry.objects.all()
    
    # Apply date filters
    if start_date:
        work_entries = work_entries.filter(work_date__gte=start_date)
    if end_date:
        work_entries = work_entries.filter(work_date__lte=end_date)
    
    # Project-wise summary
    project_stats = work_entries.values(
        'project__id',
        'project__name',
        'project__project_id',
        'project__status'
    ).annotate(
        total_hours=Sum('working_hours'),
        total_entries=Count('id'),
        unique_employees=Count('employee', distinct=True)
    ).order_by('project__name')
    
    # Calculate payment for each project
    for stat in project_stats:
        project_entries = work_entries.filter(project_id=stat['project__id'])
        employee_payments = project_entries.values('employee__man_hour_of_employee').annotate(
            hours=Sum('working_hours')
        )
        
        total_payment = sum(
            (entry['hours'] or 0) * (entry['employee__man_hour_of_employee'] or 0)
            for entry in employee_payments
        )
        stat['total_payment'] = total_payment
    
    # Overall totals
    overall_totals = work_entries.aggregate(
        total_hours=Sum('working_hours'),
        total_entries=Count('id'),
        unique_employees=Count('employee', distinct=True)
    )
    
    # Calculate overall payment
    overall_payment = sum(stat['total_payment'] for stat in project_stats)
    
    context = {
        'project_stats': project_stats,
        'overall_totals': overall_totals,
        'overall_payment': overall_payment,
        'start_date': start_date,
        'end_date': end_date,
    }
    
    return render(request, 'admin/all_projects_work_summary.html', context)

@login_required
def employee_work_report(request, employee_id):
    """View to display work report for a specific employee"""
    employee = get_object_or_404(User, id=employee_id)
    
    # Get date filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    project_id = request.GET.get('project')
    
    # Base queryset
    work_entries = WorkEntry.objects.filter(employee=employee)
    
    # Apply filters
    if start_date:
        work_entries = work_entries.filter(work_date__gte=start_date)
    if end_date:
        work_entries = work_entries.filter(work_date__lte=end_date)
    if project_id:
        work_entries = work_entries.filter(project_id=project_id)
    
    # Project-wise breakdown for this employee
    project_breakdown = work_entries.values(
        'project__id',
        'project__name',
        'project__project_id'
    ).annotate(
        total_hours=Sum('working_hours'),
        total_entries=Count('id')
    ).order_by('project__name')
    
    # Calculate payment for each project
    for breakdown in project_breakdown:
        total_hours = breakdown['total_hours'] or 0
        hourly_rate = employee.man_hour_of_employee or 0
        breakdown['total_payment'] = total_hours * hourly_rate
    
    # Employee totals
    employee_totals = work_entries.aggregate(
        total_hours=Sum('working_hours'),
        total_entries=Count('id')
    )
    
    # Total payment
    total_payment = (employee_totals['total_hours'] or 0) * (employee.man_hour_of_employee or 0)
    
    # Get projects for filter
    projects = Project.objects.filter(work_entries__employee=employee).distinct()
    
    # Recent entries
    recent_entries = work_entries.select_related('project').order_by('-work_date', '-created_at')[:15]
    
    context = {
        'employee': employee,
        'project_breakdown': project_breakdown,
        'employee_totals': employee_totals,
        'total_payment': total_payment,
        'projects': projects,
        'recent_entries': recent_entries,
        'start_date': start_date,
        'end_date': end_date,
        'selected_project': project_id,
    }
    
    return render(request, 'admin/employee_work_report.html', context)

@login_required
def export_project_report_excel(request, project_id):
    """Export project work hours report to Excel"""
    project = get_object_or_404(Project, id=project_id)
    
    # Get filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    employee_id = request.GET.get('employee')
    
    # Filter work entries
    work_entries = WorkEntry.objects.filter(project=project)
    if start_date:
        work_entries = work_entries.filter(work_date__gte=start_date)
    if end_date:
        work_entries = work_entries.filter(work_date__lte=end_date)
    if employee_id:
        work_entries = work_entries.filter(employee_id=employee_id)
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Summary sheet headers
    summary_headers = ['Employee', 'Designation', 'Total Hours', 'Hourly Rate ($)', 'Total Payment ($)', 'Entries']
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Employee statistics
    employee_stats = work_entries.values(
        'employee__first_name',
        'employee__last_name',
        'employee__designation',
        'employee__man_hour_of_employee'
    ).annotate(
        total_hours=Sum('working_hours'),
        total_entries=Count('id')
    ).order_by('employee__first_name')
    
    # Fill summary data
    row = 2
    total_hours = 0
    total_payment = 0
    
    for stat in employee_stats:
        employee_name = f"{stat['employee__first_name']} {stat['employee__last_name'] or ''}"
        hours = stat['total_hours'] or 0
        hourly_rate = stat['employee__man_hour_of_employee'] or 0
        payment = hours * hourly_rate
        
        ws_summary.cell(row=row, column=1, value=employee_name)
        ws_summary.cell(row=row, column=2, value=stat['employee__designation'] or '')
        ws_summary.cell(row=row, column=3, value=float(hours))
        ws_summary.cell(row=row, column=4, value=float(hourly_rate))
        ws_summary.cell(row=row, column=5, value=float(payment))
        ws_summary.cell(row=row, column=6, value=stat['total_entries'])
        
        total_hours += hours
        total_payment += payment
        row += 1
    
    # Add totals row
    ws_summary.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws_summary.cell(row=row, column=3, value=float(total_hours)).font = Font(bold=True)
    ws_summary.cell(row=row, column=5, value=float(total_payment)).font = Font(bold=True)
    
    # Detailed entries sheet
    ws_detail = wb.create_sheet("Detailed Entries")
    detail_headers = ['Date', 'Employee', 'Start Time', 'End Time', 'Hours', 'Hourly Rate ($)', 'Payment ($)', 'Description']
    
    for col, header in enumerate(detail_headers, 1):
        cell = ws_detail.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Fill detailed data
    detailed_entries = work_entries.select_related('employee').order_by('-work_date', '-start_time')
    
    for row, entry in enumerate(detailed_entries, 2):
        employee_name = f"{entry.employee.first_name} {entry.employee.last_name or ''}"
        hourly_rate = entry.employee.man_hour_of_employee or 0
        payment = entry.working_hours * hourly_rate
        
        ws_detail.cell(row=row, column=1, value=entry.work_date.strftime('%Y-%m-%d'))
        ws_detail.cell(row=row, column=2, value=employee_name)
        ws_detail.cell(row=row, column=3, value=entry.start_time.strftime('%H:%M'))
        ws_detail.cell(row=row, column=4, value=entry.end_time.strftime('%H:%M'))
        ws_detail.cell(row=row, column=5, value=float(entry.working_hours))
        ws_detail.cell(row=row, column=6, value=float(hourly_rate))
        ws_detail.cell(row=row, column=7, value=float(payment))
        ws_detail.cell(row=row, column=8, value=entry.description)
    
    # Auto-adjust column widths
    for ws in [ws_summary, ws_detail]:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Generate filename
    date_suffix = ""
    if start_date and end_date:
        date_suffix = f"_{start_date}_to_{end_date}"
    elif start_date:
        date_suffix = f"_from_{start_date}"
    elif end_date:
        date_suffix = f"_until_{end_date}"
    
    filename = f"project_report_{project.project_id}{date_suffix}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@login_required
def export_all_projects_excel(request):
    """Export all projects summary to Excel"""
    
    # Get filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Filter work entries
    work_entries = WorkEntry.objects.all()
    if start_date:
        work_entries = work_entries.filter(work_date__gte=start_date)
    if end_date:
        work_entries = work_entries.filter(work_date__lte=end_date)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Projects Summary"
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Headers
    headers = ['Project ID', 'Project Name', 'Status', 'Total Hours', 'Employees', 'Total Payment ($)', 'Entries']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Project statistics
    project_stats = work_entries.values(
        'project__project_id',
        'project__name',
        'project__status'
    ).annotate(
        total_hours=Sum('working_hours'),
        total_entries=Count('id'),
        unique_employees=Count('employee', distinct=True)
    ).order_by('project__name')
    
    # Fill data
    row = 2
    grand_total_hours = 0
    grand_total_payment = 0
    
    for stat in project_stats:
        # Calculate payment for this project
        project_entries = work_entries.filter(project__project_id=stat['project__project_id'])
        employee_payments = project_entries.values('employee__man_hour_of_employee').annotate(
            hours=Sum('working_hours')
        )
        
        total_payment = sum(
            (entry['hours'] or 0) * (entry['employee__man_hour_of_employee'] or 0)
            for entry in employee_payments
        )
        
        ws.cell(row=row, column=1, value=stat['project__project_id'])
        ws.cell(row=row, column=2, value=stat['project__name'])
        ws.cell(row=row, column=3, value=stat['project__status'].title())
        ws.cell(row=row, column=4, value=float(stat['total_hours'] or 0))
        ws.cell(row=row, column=5, value=stat['unique_employees'])
        ws.cell(row=row, column=6, value=float(total_payment))
        ws.cell(row=row, column=7, value=stat['total_entries'])
        
        grand_total_hours += stat['total_hours'] or 0
        grand_total_payment += total_payment
        row += 1
    
    # Add totals row
    ws.cell(row=row, column=1, value="GRAND TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=4, value=float(grand_total_hours)).font = Font(bold=True)
    ws.cell(row=row, column=6, value=float(grand_total_payment)).font = Font(bold=True)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Generate filename
    date_suffix = ""
    if start_date and end_date:
        date_suffix = f"_{start_date}_to_{end_date}"
    elif start_date:
        date_suffix = f"_from_{start_date}"
    elif end_date:
        date_suffix = f"_until_{end_date}"
    
    filename = f"all_projects_summary{date_suffix}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response



#report of all manner 

@login_required
def reports(request):
    if request.user.role != 'admin':
        return redirect('employee_dashboard')
    
    # Default date range - current month
    today = date.today()
    start_date = today.replace(day=1)
    end_date = today
    
    # Get parameters
    report_type = request.GET.get('type', 'employee')
    start_date_param = request.GET.get('start_date')
    end_date_param = request.GET.get('end_date')
    employee_id = request.GET.get('employee')
    project_id = request.GET.get('project')
    
    if start_date_param:
        start_date = date.fromisoformat(start_date_param)
    if end_date_param:
        end_date = date.fromisoformat(end_date_param)
    
    # Base query
    work_entries = WorkEntry.objects.filter(
        work_date__range=[start_date, end_date]
    ).select_related('employee', 'project')
    
    if employee_id:
        work_entries = work_entries.filter(employee_id=employee_id)
    if project_id:
        work_entries = work_entries.filter(project_id=project_id)
    
    # Generate report data
    if report_type == 'employee':
        report_data = work_entries.values('employee__email', 'employee__first_name', 'employee__last_name').annotate(
            total_hours=Sum('working_hours')
        ).order_by('employee__id')
    else:  # project
        report_data = work_entries.values('project__name').annotate(
            total_hours=Sum('working_hours')
        ).order_by('project__name')
    
    employees = User.objects.filter(role='employee', is_active=True)
    projects = Project.objects.all()
    
    context = {
        'report_data': report_data,
        'work_entries': work_entries,
        'employees': employees,
        'projects': projects,
        'report_type': report_type,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'admin/reports.html', context)



def download_reports(request):
    context = {
        'projects': Project.objects.all().order_by('name'),
        'employees': CustomUser.objects.filter(is_active=True).order_by('first_name'),
        'current_date': date.today(),
    }
    return render(request,"admin/download_reports.html", context)

@login_required
def reports_dashboard(request):
    """Main reports dashboard view"""
    context = {
        'projects': Project.objects.all().order_by('name'),
        'employees': CustomUser.objects.filter(is_active=True).order_by('first_name'),
        'current_date': date.today(),
    }
    return render(request, 'reports/reports_dashboard.html', context)

@login_required
def generate_report(request):
    """Generate and download Excel report based on parameters"""
    if request.method != 'POST':
        return redirect('reports_dashboard')
    
    # Get form parameters
    report_type = request.POST.get('report_type')  # 'employee' or 'project'
    date_range = request.POST.get('date_range')   # 'daily', 'monthly', 'custom'
    start_date = request.POST.get('start_date')
    end_date = request.POST.get('end_date')
    selected_date = request.POST.get('selected_date')
    selected_month = request.POST.get('selected_month')
    selected_year = request.POST.get('selected_year')
    employee_id = request.POST.get('employee_id')
    project_id = request.POST.get('project_id')
    
    try:
        # Determine date range
        if date_range == 'daily':
            if selected_date:
                start_date = end_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
            else:
                start_date = end_date = date.today()
        elif date_range == 'monthly':
            if selected_month and selected_year:
                year = int(selected_year)
                month = int(selected_month)
                start_date = date(year, month, 1)
                end_date = date(year, month, calendar.monthrange(year, month)[1])
            else:
                today = date.today()
                start_date = date(today.year, today.month, 1)
                end_date = date(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
        elif date_range == 'custom':
            if start_date and end_date:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            else:
                messages.error(request, 'Please provide both start and end dates for custom range.')
                return redirect('reports_dashboard')
        
        # Generate report based on type
        if report_type == 'employee':
            return generate_employee_report(start_date, end_date, employee_id)
        elif report_type == 'project':
            return generate_project_report(start_date, end_date, project_id)
        else:
            messages.error(request, 'Invalid report type selected.')
            return redirect('reports_dashboard')
            
    except Exception as e:
        messages.error(request, f'Error generating report: {str(e)}')
        return redirect('reports_dashboard')

def generate_employee_report(start_date, end_date, employee_id=None):
    """Generate employee-wise Excel report"""
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Work Report"
    
    # Set up styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Report header
    ws['A1'] = "Employee Work Report"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f"Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
    ws['A3'] = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    
    # Filter work entries
    work_entries = WorkEntry.objects.filter(
        work_date__range=[start_date, end_date]
    ).select_related('employee', 'project')
    
    if employee_id:
        work_entries = work_entries.filter(employee_id=employee_id)
    
    # Group by employee
    employees_data = {}
    for entry in work_entries:
        emp_key = entry.employee.id
        if emp_key not in employees_data:
            employees_data[emp_key] = {
                'employee': entry.employee,
                'entries': [],
                'total_hours': 0,
                'projects': set()
            }
        employees_data[emp_key]['entries'].append(entry)
        employees_data[emp_key]['total_hours'] += float(entry.working_hours)
        employees_data[emp_key]['projects'].add(entry.project.name)
    
    # Start writing data
    current_row = 5
    
    for emp_data in employees_data.values():
        employee = emp_data['employee']
        
        # Employee header
        ws[f'A{current_row}'] = f"Employee: {employee.first_name} {employee.last_name or ''}"
        ws[f'A{current_row}'].font = subheader_font
        ws[f'E{current_row}'] = f"Total Hours: {emp_data['total_hours']:.2f}"
        ws[f'E{current_row}'].font = subheader_font
        current_row += 1
        
        ws[f'A{current_row}'] = f"Designation: {employee.designation or 'N/A'}"
        ws[f'E{current_row}'] = f"Projects: {len(emp_data['projects'])}"
        current_row += 1
        
        # Table headers
        current_row += 1
        headers = ['Date', 'Project', 'Start Time', 'End Time', 'Hours', 'Description']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        current_row += 1
        
        # Employee work entries
        for entry in sorted(emp_data['entries'], key=lambda x: x.work_date):
            ws[f'A{current_row}'] = entry.work_date.strftime('%Y-%m-%d')
            ws[f'B{current_row}'] = entry.project.name
            ws[f'C{current_row}'] = entry.start_time.strftime('%H:%M')
            ws[f'D{current_row}'] = entry.end_time.strftime('%H:%M')
            ws[f'E{current_row}'] = float(entry.working_hours)
            ws[f'F{current_row}'] = entry.description
            
            # Apply borders
            for col in range(1, 7):
                ws.cell(row=current_row, column=col).border = border
            
            current_row += 1
        
        current_row += 2  # Space between employees
    
    # Adjust column widths
    column_widths = [15, 25, 12, 12, 10, 40]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"employee_report_{start_date}_{end_date}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

def generate_project_report(start_date, end_date, project_id=None):
    """Generate project-wise Excel report"""
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project Work Report"
    
    # Set up styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Report header
    ws['A1'] = "Project Work Report"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f"Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
    ws['A3'] = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    
    # Filter work entries
    work_entries = WorkEntry.objects.filter(
        work_date__range=[start_date, end_date]
    ).select_related('employee', 'project')
    
    if project_id:
        work_entries = work_entries.filter(project_id=project_id)
    
    # Group by project
    projects_data = {}
    for entry in work_entries:
        proj_key = entry.project.id
        if proj_key not in projects_data:
            projects_data[proj_key] = {
                'project': entry.project,
                'entries': [],
                'total_hours': 0,
                'employees': set()
            }
        projects_data[proj_key]['entries'].append(entry)
        projects_data[proj_key]['total_hours'] += float(entry.working_hours)
        projects_data[proj_key]['employees'].add(f"{entry.employee.first_name} {entry.employee.last_name or ''}")
    
    # Start writing data
    current_row = 5
    
    for proj_data in projects_data.values():
        project = proj_data['project']
        
        # Project header
        ws[f'A{current_row}'] = f"Project: {project.name} ({project.project_id})"
        ws[f'A{current_row}'].font = subheader_font
        ws[f'E{current_row}'] = f"Total Hours: {proj_data['total_hours']:.2f}"
        ws[f'E{current_row}'].font = subheader_font
        current_row += 1
        
        ws[f'A{current_row}'] = f"Status: {project.status.title()}"
        ws[f'E{current_row}'] = f"Team Members: {len(proj_data['employees'])}"
        current_row += 1
        
        # Table headers
        current_row += 1
        headers = ['Date', 'Employee', 'Start Time', 'End Time', 'Hours', 'Description']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        current_row += 1
        
        # Project work entries
        for entry in sorted(proj_data['entries'], key=lambda x: x.work_date):
            ws[f'A{current_row}'] = entry.work_date.strftime('%Y-%m-%d')
            ws[f'B{current_row}'] = f"{entry.employee.first_name} {entry.employee.last_name or ''}"
            ws[f'C{current_row}'] = entry.start_time.strftime('%H:%M')
            ws[f'D{current_row}'] = entry.end_time.strftime('%H:%M')
            ws[f'E{current_row}'] = float(entry.working_hours)
            ws[f'F{current_row}'] = entry.description
            
            # Apply borders
            for col in range(1, 7):
                ws.cell(row=current_row, column=col).border = border
            
            current_row += 1
        
        current_row += 2  # Space between projects
    
    # Adjust column widths
    column_widths = [15, 25, 12, 12, 10, 40]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"project_report_{start_date}_{end_date}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

@login_required
def get_report_summary(request):
    """AJAX endpoint to get report summary data"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        # Get parameters
        report_type = request.POST.get('report_type')
        date_range = request.POST.get('date_range')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        selected_date = request.POST.get('selected_date')
        selected_month = request.POST.get('selected_month')
        selected_year = request.POST.get('selected_year')
        employee_id = request.POST.get('employee_id')
        project_id = request.POST.get('project_id')
        
        # Determine date range
        if date_range == 'daily':
            if selected_date:
                start_date = end_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
            else:
                start_date = end_date = date.today()
        elif date_range == 'monthly':
            if selected_month and selected_year:
                year = int(selected_year)
                month = int(selected_month)
                start_date = date(year, month, 1)
                end_date = date(year, month, calendar.monthrange(year, month)[1])
            else:
                today = date.today()
                start_date = date(today.year, today.month, 1)
                end_date = date(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
        elif date_range == 'custom':
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # Filter work entries
        work_entries = WorkEntry.objects.filter(
            work_date__range=[start_date, end_date]
        ).select_related('employee', 'project')
        
        if employee_id:
            work_entries = work_entries.filter(employee_id=employee_id)
        if project_id:
            work_entries = work_entries.filter(project_id=project_id)
        
        # Calculate summary
        total_hours = work_entries.aggregate(total=Sum('working_hours'))['total'] or 0
        total_entries = work_entries.count()
        unique_employees = work_entries.values('employee').distinct().count()
        unique_projects = work_entries.values('project').distinct().count()
        
        return JsonResponse({
            'success': True,
            'summary': {
                'total_hours': float(total_hours),
                'total_entries': total_entries,
                'unique_employees': unique_employees,
                'unique_projects': unique_projects,
                'date_range': f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
            }
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)